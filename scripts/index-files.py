#!/usr/bin/env python3
"""Index parseable files from synced personal data into Qdrant personal_data.

Scans /opt/athanor/personal-data/ (or specified path) for documents,
parses them, chunks, embeds, and upserts to Qdrant. Supports incremental
indexing via content hash comparison.

Usage:
    python3 scripts/index-files.py                        # full index (Node 1)
    python3 scripts/index-files.py --path /path/to/dir    # custom path
    python3 scripts/index-files.py --dry-run              # preview only
    python3 scripts/index-files.py --force                # re-index all
    python3 scripts/index-files.py --category work        # filter by path

Options:
    --path PATH       Directory to scan (default: auto-detect Node 1 or local)
    --dry-run         Show what would be indexed
    --force           Re-index even if content hash matches
    --category CAT    Only index files whose path contains CAT
    -h, --help        Show this help

Runs on DEV (via SSH) or directly on Node 1 inside the container.
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
import urllib.request
from pathlib import Path

# --- Config ---
QDRANT_URL = "http://192.168.1.244:6333"
LITELLM_URL = (os.environ.get("ATHANOR_LITELLM_URL") or "http://192.168.1.203:4000").rstrip("/") + "/v1"
LITELLM_KEY = (
    os.environ.get("ATHANOR_LITELLM_API_KEY")
    or os.environ.get("LITELLM_API_KEY")
    or os.environ.get("OPENAI_API_KEY", "")
)
COLLECTION = "personal_data"

# Auto-detect: Node 1 container path or DEV via SSH
DEFAULT_PATHS = [
    "/data/personal",               # Inside agent container
    "/opt/athanor/personal-data",   # Node 1 host
]

PARSEABLE_EXTENSIONS = {
    ".md", ".txt", ".json", ".csv", ".yaml", ".yml", ".toml",
    ".py", ".js", ".ts", ".tsx", ".jsx", ".sh", ".bash",
    ".html", ".htm", ".xml", ".cfg", ".ini", ".conf",
    ".pdf", ".docx", ".xlsx", ".xls",
}

SKIP_PATTERNS = [
    "node_modules", "__pycache__", ".git", ".next",
    "Thumbs.db", "desktop.ini", ".DS_Store",
    "package-lock.json", "ventoy",
]

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
CHUNK_SIZE = 1500
CHUNK_OVERLAP = 200
BATCH_SIZE = 20


# --- File parsing ---

def parse_text_file(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def parse_xlsx(filepath: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(max_row=500, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        wb.close()
        return "\n".join(lines)
    except ImportError:
        return f"[XLSX parsing unavailable — install openpyxl] {filepath}"
    except Exception as e:
        return f"[XLSX parse error: {e}]"


def parse_pdf(filepath: str) -> str:
    try:
        import pymupdf
        doc = pymupdf.open(filepath)
        pages = [page.get_text() for page in doc]
        doc.close()
        return "\n\n".join(pages)
    except ImportError:
        return f"[PDF parsing unavailable — install pymupdf]"
    except Exception as e:
        return f"[PDF parse error: {e}]"


def parse_docx(filepath: str) -> str:
    try:
        import docx
        doc = docx.Document(filepath)
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return f"[DOCX parsing unavailable — install python-docx]"
    except Exception as e:
        return f"[DOCX parse error: {e}]"


def parse_file(filepath: str) -> str:
    ext = Path(filepath).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return parse_xlsx(filepath)
    elif ext == ".pdf":
        return parse_pdf(filepath)
    elif ext == ".docx":
        return parse_docx(filepath)
    else:
        return parse_text_file(filepath)


# --- Chunking ---

def chunk_text(text: str) -> list[str]:
    if len(text) <= CHUNK_SIZE:
        return [text]

    chunks = []
    start = 0
    while start < len(text):
        end = start + CHUNK_SIZE
        if end < len(text):
            # Try to break at paragraph, then sentence, then newline
            para = text.rfind("\n\n", start + CHUNK_SIZE // 2, end)
            if para > start:
                end = para
            else:
                for sep in [". ", ".\n", "\n"]:
                    sent = text.rfind(sep, start + CHUNK_SIZE // 2, end)
                    if sent > start:
                        end = sent + len(sep)
                        break
        chunks.append(text[start:end].strip())
        start = end - CHUNK_OVERLAP

    return [c for c in chunks if len(c) > 50]


# --- Classify ---

def classify_path(filepath: str) -> tuple[str, str]:
    """Return (category, subcategory) based on file path."""
    path_lower = filepath.lower()

    if "athanor-reference" in path_lower:
        return "file", "athanor_reference"
    if "sharex" in path_lower or "configs" in path_lower:
        return "file", "configs"
    if "bki" in path_lower or "tracker" in path_lower:
        return "file", "work_energy"
    if any(x in path_lower for x in ["carver", "gladstone", "testing", "volta"]):
        return "file", "work_energy"
    if "energy" in path_lower or "audit" in path_lower:
        return "file", "work_energy"
    if "finance" in path_lower or "debt" in path_lower:
        return "file", "finance"
    if "athanor" in path_lower:
        return "file", "athanor_docs"
    if "claude" in path_lower:
        return "file", "claude_docs"
    if any(x in path_lower for x in ["photo", "jpg", "jpeg", "png"]):
        return "file", "photos"
    if "download" in path_lower:
        return "file", "downloads"
    if "document" in path_lower:
        return "file", "documents"

    return "file", "general"


# --- Qdrant + Embedding ---

def get_embedding(text: str) -> list[float]:
    body = json.dumps({
        "model": "embedding",
        "input": text[:2000],
    }).encode()
    req = urllib.request.Request(
        f"{LITELLM_URL}/embeddings",
        data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {LITELLM_KEY}",
        },
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read())
    return data["data"][0]["embedding"]


def get_existing_hashes() -> dict[str, str]:
    """Get content_hash for all file-category points in Qdrant."""
    hashes = {}
    offset = None
    while True:
        body: dict = {
            "limit": 100,
            "with_payload": {"include": ["source", "content_hash"]},
            "with_vector": False,
            "filter": {"must": [{"key": "category", "match": {"value": "file"}}]},
        }
        if offset is not None:
            body["offset"] = offset

        payload = json.dumps(body).encode()
        req = urllib.request.Request(
            f"{QDRANT_URL}/collections/{COLLECTION}/points/scroll",
            data=payload,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())

        result = data.get("result", {})
        points = result.get("points", [])
        for p in points:
            pl = p.get("payload", {})
            source = pl.get("source", "")
            content_hash = pl.get("content_hash", "")
            if source and content_hash:
                hashes[source] = content_hash

        next_offset = result.get("next_page_offset")
        if not next_offset or not points:
            break
        offset = next_offset

    return hashes


def upsert_batch(points: list[dict]):
    payload = json.dumps({"points": points}).encode()
    req = urllib.request.Request(
        f"{QDRANT_URL}/collections/{COLLECTION}/points",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="PUT",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        if resp.status not in (200, 201):
            raise RuntimeError(f"Qdrant upsert failed: HTTP {resp.status}")


# --- Main ---

def find_scan_path() -> str | None:
    for p in DEFAULT_PATHS:
        if os.path.isdir(p):
            return p
    return None


def main():
    parser = argparse.ArgumentParser(description="Index synced personal files into Qdrant")
    parser.add_argument("--path", type=str, default="", help="Directory to scan")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    parser.add_argument("--force", action="store_true", help="Re-index all files")
    parser.add_argument("--category", type=str, default="", help="Filter by path substring")
    args = parser.parse_args()

    scan_path = args.path or find_scan_path()
    if not scan_path or not os.path.isdir(scan_path):
        print(f"ERROR: No scan path found. Use --path or run on Node 1.", file=sys.stderr)
        sys.exit(1)

    print(f"=== Athanor File Indexer ===\n", file=sys.stderr)
    print(f"Scan path: {scan_path}", file=sys.stderr)

    # Find parseable files
    files = []
    for root, dirs, filenames in os.walk(scan_path):
        # Skip unwanted directories
        dirs[:] = [d for d in dirs if d not in SKIP_PATTERNS]

        for fname in filenames:
            ext = Path(fname).suffix.lower()
            if ext not in PARSEABLE_EXTENSIONS:
                continue

            filepath = os.path.join(root, fname)
            rel_path = os.path.relpath(filepath, scan_path)

            if any(s in filepath for s in SKIP_PATTERNS):
                continue
            if args.category and args.category.lower() not in filepath.lower():
                continue

            try:
                size = os.path.getsize(filepath)
            except OSError:
                continue

            if size > MAX_FILE_SIZE or size == 0:
                continue

            files.append((filepath, rel_path, size))

    print(f"Found {len(files)} parseable files\n", file=sys.stderr)

    if not files:
        print("Nothing to index.", file=sys.stderr)
        return 0

    if args.dry_run:
        for fp, rel, size in files:
            cat, sub = classify_path(fp)
            size_str = f"{size / 1024:.0f} KB" if size < 1024 * 1024 else f"{size / (1024 * 1024):.1f} MB"
            print(f"  {rel}  ({size_str}, {sub})", file=sys.stderr)
        print(f"\nDry run: {len(files)} files would be indexed.", file=sys.stderr)
        return 0

    # Get existing hashes for incremental indexing
    existing_hashes = {} if args.force else get_existing_hashes()
    print(f"Existing file points: {len(existing_hashes)}", file=sys.stderr)

    indexed = 0
    skipped = 0
    errors = 0
    batch = []

    for i, (filepath, rel_path, size) in enumerate(files):
        source = f"file:{rel_path}"
        print(f"  [{i + 1}/{len(files)}] {rel_path}...", end="", file=sys.stderr)

        try:
            content = parse_file(filepath)
            if len(content.strip()) < 50:
                print(" (too short)", file=sys.stderr)
                skipped += 1
                continue

            content_hash = hashlib.md5(content.encode()).hexdigest()

            # Skip if hash matches existing
            if source in existing_hashes and existing_hashes[source] == content_hash:
                print(" (unchanged)", file=sys.stderr)
                skipped += 1
                continue

            category, subcategory = classify_path(filepath)
            chunks = chunk_text(content)

            for ci, chunk in enumerate(chunks):
                vector = get_embedding(chunk)
                point_id = hashlib.md5(f"{source}:chunk:{ci}".encode()).hexdigest()

                batch.append({
                    "id": point_id,
                    "vector": vector,
                    "payload": {
                        "source": source,
                        "filename": os.path.basename(filepath),
                        "title": Path(filepath).stem.replace("_", " ").replace("-", " "),
                        "category": category,
                        "subcategory": subcategory,
                        "text": chunk[:2000],
                        "content_hash": content_hash,
                        "chunk_index": ci,
                        "total_chunks": len(chunks),
                        "file_size": size,
                        "file_modified": time.strftime(
                            "%Y-%m-%dT%H:%M:%S",
                            time.localtime(os.path.getmtime(filepath)),
                        ),
                        "indexed_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    },
                })

            if len(batch) >= BATCH_SIZE:
                upsert_batch(batch)
                batch = []

            print(f" {len(chunks)} chunks ({subcategory})", file=sys.stderr)
            indexed += 1

        except Exception as e:
            print(f" ERROR: {e}", file=sys.stderr)
            errors += 1

    # Flush remaining
    if batch:
        upsert_batch(batch)

    print(f"\n{'=' * 50}", file=sys.stderr)
    print(f"Indexed:  {indexed} files", file=sys.stderr)
    print(f"Skipped:  {skipped} (unchanged or too short)", file=sys.stderr)
    print(f"Errors:   {errors}", file=sys.stderr)
    print(f"Total:    {len(files)} files scanned", file=sys.stderr)

    # Get updated collection count
    try:
        req = urllib.request.Request(f"{QDRANT_URL}/collections/{COLLECTION}")
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
        count = data.get("result", {}).get("points_count", "?")
        print(f"\nQdrant {COLLECTION}: {count} total points", file=sys.stderr)
    except Exception:
        pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
