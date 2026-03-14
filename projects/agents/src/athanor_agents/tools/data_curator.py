"""Data Curator tools — scan, parse, analyze, and index personal data."""

import hashlib
import json
import logging
import os
import time
from pathlib import Path

import httpx
from langchain_core.tools import tool

from ..config import settings
from ..services import registry

_QDRANT_URL = settings.qdrant_url
_EMBEDDING_URL = registry.litellm_openai_url
_EMBEDDING_KEY = settings.litellm_api_key
_LLM_URL = settings.llm_base_url
_LLM_KEY = settings.litellm_api_key

# Scannable root directories (inside the container, mapped via docker volumes)
# /data/personal maps to VAULT NFS personal data sync target
# /data/local maps to WSL /mnt/c and /mnt/d via bind mount
SCAN_ROOTS = {
    "personal": "/data/personal",
    "local_c": "/data/local/c",
    "local_d": "/data/local/d",
    "gdrive": "/data/personal/gdrive",
}

COLLECTION = "personal_data"

# File extensions we can parse
PARSEABLE_EXTENSIONS = {
    # Text-based
    ".md", ".txt", ".json", ".csv", ".yaml", ".yml", ".toml",
    ".py", ".js", ".ts", ".tsx", ".jsx", ".sh", ".bash",
    ".html", ".htm", ".xml", ".cfg", ".ini", ".conf", ".log",
    # Documents (require extra parsing)
    ".pdf", ".docx", ".xlsx",
}

# Max file size to parse (10 MB)
MAX_FILE_SIZE = 10 * 1024 * 1024


def _get_embedding(text: str) -> list[float]:
    """Get embedding vector from the embedding model via LiteLLM."""
    resp = httpx.post(
        f"{_EMBEDDING_URL}/embeddings",
        json={"model": "embedding", "input": text[:2000]},
        headers={"Authorization": f"Bearer {_EMBEDDING_KEY}"},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["data"][0]["embedding"]


def _text_to_id(text: str) -> str:
    """Generate a deterministic point ID from text content."""
    return hashlib.md5(text.encode()).hexdigest()


def _chunk_text(text: str, chunk_size: int = 1500, overlap: int = 200) -> list[str]:
    """Split text into overlapping chunks at paragraph/sentence boundaries."""
    if len(text) <= chunk_size:
        return [text]

    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size

        if end < len(text):
            para_break = text.rfind("\n\n", start + chunk_size // 2, end)
            if para_break > start:
                end = para_break
            else:
                for sep in [". ", ".\n", "\n"]:
                    sent_break = text.rfind(sep, start + chunk_size // 2, end)
                    if sent_break > start:
                        end = sent_break + len(sep)
                        break

        chunks.append(text[start:end].strip())
        start = end - overlap

    return [c for c in chunks if len(c) > 50]


def _parse_text_file(filepath: str) -> str:
    """Read a plain text file."""
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _parse_pdf(filepath: str) -> str:
    """Parse PDF using pymupdf if available, fallback to basic read."""
    try:
        import pymupdf
        doc = pymupdf.open(filepath)
        pages = []
        for page in doc:
            pages.append(page.get_text())
        doc.close()
        return "\n\n".join(pages)
    except ImportError:
        return f"[PDF parsing unavailable — install pymupdf] {filepath}"
    except Exception as e:
        return f"[PDF parse error: {e}]"


def _parse_docx(filepath: str) -> str:
    """Parse Word document using python-docx if available."""
    try:
        import docx
        doc = docx.Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        return f"[DOCX parsing unavailable — install python-docx] {filepath}"
    except Exception as e:
        return f"[DOCX parse error: {e}]"


def _parse_xlsx(filepath: str) -> str:
    """Parse Excel spreadsheet using openpyxl if available."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(max_row=500, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        wb.close()
        return "\n".join(lines)
    except ImportError:
        return f"[XLSX parsing unavailable — install openpyxl] {filepath}"
    except Exception as e:
        return f"[XLSX parse error: {e}]"


def _parse_file(filepath: str) -> str:
    """Parse a file based on its extension."""
    ext = Path(filepath).suffix.lower()

    if ext == ".pdf":
        return _parse_pdf(filepath)
    elif ext == ".docx":
        return _parse_docx(filepath)
    elif ext in (".xlsx", ".xls"):
        return _parse_xlsx(filepath)
    else:
        return _parse_text_file(filepath)


def _ensure_collection():
    """Ensure the personal_data Qdrant collection exists."""
    try:
        resp = httpx.get(f"{_QDRANT_URL}/collections/{COLLECTION}", timeout=5)
        if resp.status_code == 200:
            return
    except Exception as e:
        logging.getLogger(__name__).debug("Collection check failed: %s", e)

    httpx.put(
        f"{_QDRANT_URL}/collections/{COLLECTION}",
        json={
            "vectors": {"size": 1024, "distance": "Cosine"},
        },
        timeout=10,
    )


def _classify_file(filepath: str) -> str:
    """Classify a file into a category based on its path."""
    path_lower = filepath.lower()

    if "finance" in path_lower or "debt" in path_lower or "budget" in path_lower:
        return "finance"
    if "energy" in path_lower or "audit" in path_lower or "bki" in path_lower or "duct" in path_lower:
        return "work_energy"
    if any(x in path_lower for x in ["athanor", "hydra", "kaizen", "ai-dev", "system-bible"]):
        return "ai_system"
    if any(x in path_lower for x in ["facebook", "social", "online data"]):
        return "social_media"
    if any(x in path_lower for x in ["comfyui", "swarmui", "models", "huggingface"]):
        return "ai_models"
    if any(x in path_lower for x in ["performer", "stash", "nsfw"]):
        return "adult"
    if any(x in path_lower for x in ["sovereign", "master list"]):
        return "personal_archive"
    if "chatgpt" in path_lower or "claude" in path_lower:
        return "ai_conversations"
    if any(x in path_lower for x in [".py", ".js", ".ts", ".sh"]):
        return "code"
    if any(x in path_lower for x in [".pdf", ".docx", ".xlsx"]):
        return "document"

    return "general"


@tool
def scan_directory(
    root: str = "personal",
    path: str = "",
    extensions: str = "",
    max_depth: int = 5,
) -> str:
    """Scan a directory and list all files with their sizes and types.

    Use this to discover what personal data exists before parsing/indexing.
    Scans recursively up to max_depth levels.

    Available roots:
    - "personal" — VAULT NFS personal data (/data/personal)
    - "local_c" — Windows C: drive (/data/local/c)
    - "local_d" — Windows D: drive (/data/local/d)
    - "gdrive" — Google Drive sync target (/data/personal/gdrive)

    Args:
        root: Which root directory to scan (personal, local_c, local_d, gdrive)
        path: Subdirectory path within the root (e.g. "Documents/Work")
        extensions: Comma-separated file extensions to filter (e.g. ".pdf,.xlsx"). Empty = all parseable.
        max_depth: Maximum directory depth to scan (default 5)
    """
    base = SCAN_ROOTS.get(root)
    if not base:
        return f"Unknown root '{root}'. Available: {', '.join(SCAN_ROOTS.keys())}"

    scan_path = Path(base) / path if path else Path(base)
    if not scan_path.exists():
        return f"Path does not exist: {scan_path}"

    ext_filter = set()
    if extensions:
        ext_filter = {e.strip().lower() if e.startswith(".") else f".{e.strip().lower()}"
                      for e in extensions.split(",")}

    files = []
    dirs = []
    total_size = 0

    try:
        for item in sorted(scan_path.rglob("*")):
            # Respect depth limit
            rel = item.relative_to(scan_path)
            if len(rel.parts) > max_depth:
                continue

            if item.is_dir():
                dirs.append(str(rel))
            elif item.is_file():
                ext = item.suffix.lower()
                if ext_filter and ext not in ext_filter:
                    continue
                if not ext_filter and ext not in PARSEABLE_EXTENSIONS:
                    continue

                size = item.stat().st_size
                total_size += size
                modified = time.strftime(
                    "%Y-%m-%d", time.localtime(item.stat().st_mtime)
                )
                category = _classify_file(str(item))
                size_str = (
                    f"{size / 1024:.0f} KB" if size < 1024 * 1024
                    else f"{size / (1024 * 1024):.1f} MB"
                )
                files.append(f"  {rel}  ({size_str}, {modified}, {category})")

    except PermissionError:
        return f"Permission denied scanning: {scan_path}"

    lines = [f"Scan of {root}/{path or '.'}: {len(files)} files, {len(dirs)} directories"]
    lines.append(f"Total parseable size: {total_size / (1024 * 1024):.1f} MB")
    lines.append("")

    if dirs and len(dirs) <= 50:
        lines.append("Directories:")
        for d in dirs[:50]:
            lines.append(f"  {d}/")
        lines.append("")

    lines.append("Files:")
    for f in files[:200]:
        lines.append(f)

    if len(files) > 200:
        lines.append(f"  ... and {len(files) - 200} more files")

    return "\n".join(lines)


@tool
def parse_document(filepath: str, root: str = "personal") -> str:
    """Parse a single file and return its text content.

    Use this to read and understand a specific file before deciding whether to index it.
    Supports: .md, .txt, .json, .csv, .py, .js, .html, .pdf, .docx, .xlsx

    Args:
        filepath: Path relative to the root directory
        root: Which root directory (personal, local_c, local_d, gdrive)
    """
    base = SCAN_ROOTS.get(root)
    if not base:
        return f"Unknown root '{root}'. Available: {', '.join(SCAN_ROOTS.keys())}"

    full_path = Path(base) / filepath
    if not full_path.exists():
        return f"File not found: {full_path}"
    if not full_path.is_file():
        return f"Not a file: {full_path}"
    if full_path.stat().st_size > MAX_FILE_SIZE:
        return f"File too large ({full_path.stat().st_size / (1024*1024):.1f} MB > 10 MB limit)"

    try:
        content = _parse_file(str(full_path))
        # Truncate very long files for display
        if len(content) > 8000:
            return content[:8000] + f"\n\n[... truncated, {len(content)} total chars]"
        return content
    except Exception as e:
        return f"Parse error: {e}"


@tool
def analyze_content(text: str, instruction: str = "classify and summarize") -> str:
    """Use the local LLM to analyze text content.

    Sends the text to the local LLM for analysis — classification, summarization,
    entity extraction, or any other analysis task. Zero API cost.

    Args:
        text: The text content to analyze (will be truncated to 4000 chars)
        instruction: What to do with the text (e.g. "classify and summarize",
                    "extract entities and relationships", "identify key topics")
    """
    truncated = text[:4000]

    try:
        resp = httpx.post(
            f"{_LLM_URL}/chat/completions",
            json={
                "model": "reasoning",
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "You are a data analysis assistant. Analyze the provided text "
                            "and follow the instruction precisely. Be concise and structured. "
                            "Output in English only."
                        ),
                    },
                    {
                        "role": "user",
                        "content": f"Instruction: {instruction}\n\nText:\n{truncated}",
                    },
                ],
                "temperature": 0.3,
                "max_tokens": 1000,
            },
            headers={"Authorization": f"Bearer {_LLM_KEY}"},
            timeout=60,
        )
        resp.raise_for_status()
        result = resp.json()["choices"][0]["message"]["content"]

        # Strip <think> blocks if present (Qwen3 reasoning)
        import re
        result = re.sub(r"<think>.*?</think>", "", result, flags=re.DOTALL).strip()
        return result
    except Exception as e:
        return f"Analysis error: {e}"


@tool
def index_document(
    filepath: str,
    root: str = "personal",
    category: str = "",
    tags: str = "",
) -> str:
    """Parse, chunk, embed, and index a file into the personal_data Qdrant collection.

    This makes the file searchable by all agents. Each chunk gets its own vector
    with metadata (source path, category, tags, timestamp).

    Args:
        filepath: Path relative to the root directory
        root: Which root directory (personal, local_c, local_d, gdrive)
        category: Override auto-detected category (e.g. "finance", "work_energy", "ai_system")
        tags: Comma-separated tags (e.g. "bki,energy-audit,february-2026")
    """
    base = SCAN_ROOTS.get(root)
    if not base:
        return f"Unknown root '{root}'. Available: {', '.join(SCAN_ROOTS.keys())}"

    full_path = Path(base) / filepath
    if not full_path.exists():
        return f"File not found: {full_path}"
    if full_path.stat().st_size > MAX_FILE_SIZE:
        return f"File too large: {full_path.stat().st_size / (1024*1024):.1f} MB"

    _ensure_collection()

    try:
        content = _parse_file(str(full_path))
        if len(content.strip()) < 50:
            return f"File too short to index: {filepath}"

        auto_category = category or _classify_file(str(full_path))
        tag_list = [t.strip() for t in tags.split(",") if t.strip()] if tags else []
        source = f"{root}:{filepath}"
        content_hash = hashlib.md5(content.encode()).hexdigest()

        chunks = _chunk_text(content)

        # Embed and upsert in batches
        points = []
        for i, chunk in enumerate(chunks):
            try:
                vector = _get_embedding(chunk)
                point_id = _text_to_id(f"{source}:chunk:{i}")

                points.append({
                    "id": point_id,
                    "vector": vector,
                    "payload": {
                        "source": source,
                        "filename": full_path.name,
                        "category": auto_category,
                        "tags": tag_list,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "text": chunk[:2000],
                        "content_hash": content_hash,
                        "file_size": full_path.stat().st_size,
                        "file_modified": time.strftime(
                            "%Y-%m-%dT%H:%M:%S",
                            time.localtime(full_path.stat().st_mtime),
                        ),
                        "indexed_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    },
                })
            except Exception as e:
                return f"Embedding failed on chunk {i}: {e}"

        # Upsert to Qdrant
        if points:
            resp = httpx.put(
                f"{_QDRANT_URL}/collections/{COLLECTION}/points",
                json={"points": points},
                timeout=30,
            )
            resp.raise_for_status()

        return (
            f"Indexed {filepath}: {len(chunks)} chunks, category={auto_category}, "
            f"tags={tag_list}, hash={content_hash[:8]}"
        )
    except Exception as e:
        return f"Index error: {e}"


@tool
def search_personal(query: str, category: str = "", limit: int = 5) -> str:
    """Search indexed personal data using semantic vector search.

    Finds relevant content across all indexed personal files by meaning.

    Args:
        query: Natural language search query
        category: Filter by category (finance, work_energy, ai_system, social_media, etc). Empty = all.
        limit: Number of results (default 5)
    """
    _ensure_collection()

    try:
        vector = _get_embedding(query)
        body: dict = {
            "vector": vector,
            "limit": limit,
            "with_payload": True,
            "score_threshold": 0.25,
        }
        if category:
            body["filter"] = {
                "must": [{"key": "category", "match": {"value": category}}]
            }

        resp = httpx.post(
            f"{_QDRANT_URL}/collections/{COLLECTION}/points/search",
            json=body,
            timeout=10,
        )
        resp.raise_for_status()
        results = resp.json().get("result", [])

        if not results:
            return f"No personal data matches for: {query}"

        lines = []
        for i, r in enumerate(results, 1):
            score = r.get("score", 0)
            payload = r.get("payload", {})
            source = payload.get("source", "unknown")
            category_val = payload.get("category", "")
            tags = payload.get("tags", [])
            text = payload.get("text", "")[:400]
            chunk_idx = payload.get("chunk_index", 0)
            total = payload.get("total_chunks", 1)
            modified = payload.get("file_modified", "")[:10]

            lines.append(f"[{i}] {source} (score: {score:.3f})")
            lines.append(f"    Category: {category_val} | Chunk {chunk_idx+1}/{total} | Modified: {modified}")
            if tags:
                lines.append(f"    Tags: {', '.join(tags)}")
            lines.append(f"    {text}")
            lines.append("")
        return "\n".join(lines)
    except Exception as e:
        return f"Search error: {e}"


@tool
def get_scan_status() -> str:
    """Get statistics about the personal data collection — total indexed, by category, freshness.

    Use this to understand what has been indexed and what gaps exist.
    """
    _ensure_collection()

    try:
        resp = httpx.get(f"{_QDRANT_URL}/collections/{COLLECTION}", timeout=5)
        resp.raise_for_status()
        info = resp.json().get("result", {})
        total = info.get("points_count", 0)

        lines = [f"# Personal Data Index Status"]
        lines.append(f"Total indexed chunks: {total}")
        lines.append("")

        # Get category breakdown via scroll
        categories: dict[str, int] = {}
        sources: set[str] = set()
        offset = None
        batch_count = 0

        while batch_count < 10:  # Safety limit
            body: dict = {
                "limit": 100,
                "with_payload": ["category", "source"],
            }
            if offset:
                body["offset"] = offset

            resp = httpx.post(
                f"{_QDRANT_URL}/collections/{COLLECTION}/points/scroll",
                json=body,
                timeout=10,
            )
            resp.raise_for_status()
            result = resp.json().get("result", {})
            points = result.get("points", [])
            next_offset = result.get("next_page_offset")

            for p in points:
                payload = p.get("payload", {})
                cat = payload.get("category", "uncategorized")
                src = payload.get("source", "")
                categories[cat] = categories.get(cat, 0) + 1
                if src:
                    sources.add(src.split(":chunk:")[0] if ":chunk:" in src else src)

            if not next_offset or not points:
                break
            offset = next_offset
            batch_count += 1

        if categories:
            lines.append("Categories:")
            for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
                lines.append(f"  {cat}: {count} chunks")

        lines.append(f"\nUnique files indexed: {len(sources)}")

        # Check which roots are accessible
        lines.append("\nRoot accessibility:")
        for name, path in SCAN_ROOTS.items():
            exists = os.path.exists(path)
            lines.append(f"  {name} ({path}): {'accessible' if exists else 'NOT MOUNTED'}")

        return "\n".join(lines)
    except Exception as e:
        return f"Status error: {e}"


@tool
def sync_gdrive() -> str:
    """Trigger a Google Drive sync via rclone.

    Syncs the configured Google Drive remote to /data/personal/gdrive/.
    Requires rclone to be configured with Google Drive OAuth (one-time setup by Shaun).

    Returns sync output or error if rclone is not configured.
    """
    import subprocess

    rclone_path = "/usr/local/bin/rclone"
    if not os.path.exists(rclone_path):
        rclone_path = "rclone"

    try:
        # Check if gdrive remote is configured
        result = subprocess.run(
            [rclone_path, "listremotes"],
            capture_output=True, text=True, timeout=10,
        )
        remotes = result.stdout.strip().split("\n")

        gdrive_remote = None
        for remote in remotes:
            if "gdrive" in remote.lower() or "google" in remote.lower():
                gdrive_remote = remote.rstrip(":")
                break

        if not gdrive_remote:
            return (
                "No Google Drive remote configured in rclone.\n"
                "Shaun needs to run: rclone config\n"
                "Choose 'New remote' → 'Google Drive' → complete OAuth in browser."
            )

        # Run sync
        sync_target = SCAN_ROOTS.get("gdrive", "/data/personal/gdrive")
        os.makedirs(sync_target, exist_ok=True)

        result = subprocess.run(
            [
                rclone_path, "sync",
                f"{gdrive_remote}:", sync_target,
                "--transfers", "4",
                "--stats", "0",
                "-v",
            ],
            capture_output=True, text=True, timeout=600,
        )

        if result.returncode == 0:
            return f"Google Drive sync complete to {sync_target}\n{result.stdout[-500:]}"
        else:
            return f"Sync failed (exit {result.returncode}):\n{result.stderr[-500:]}"

    except FileNotFoundError:
        return "rclone not found. Install it first."
    except subprocess.TimeoutExpired:
        return "Sync timed out after 10 minutes. Files may be partially synced."
    except Exception as e:
        return f"Sync error: {e}"


DATA_CURATOR_TOOLS = [
    scan_directory,
    parse_document,
    analyze_content,
    index_document,
    search_personal,
    get_scan_status,
    sync_gdrive,
]
